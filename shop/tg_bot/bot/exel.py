import os
import openpyxl


def create_excel_if_not_exists(file_path):
    if not os.path.exists(file_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Payments"

        headers = ["User_id", "Товар", "Сумма", "Номер", "Адресс"]

        for col_num, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_num, value=header)

        wb.save(file_path)
        print(f"Создан новый файл Excel с заголовками: {file_path}")


def add_payment_to_excel(BASE_DIR, user_id, item, amount, number, adress):
    data_dir = os.path.join(BASE_DIR, 'data')
    if not os.path.exists(data_dir):
        os.mkdir(data_dir)

    file_path = os.path.join(data_dir, 'payments.xlsx')

    create_excel_if_not_exists(file_path)

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    next_row = ws.max_row + 1

    ws.cell(row=next_row, column=1, value=user_id)
    ws.cell(row=next_row, column=2, value=item)
    ws.cell(row=next_row, column=3, value=amount)
    ws.cell(row=next_row, column=4, value=number)
    ws.cell(row=next_row, column=4, value=adress)

    wb.save(file_path)
